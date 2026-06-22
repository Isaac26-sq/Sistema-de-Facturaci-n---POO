import io
from datetime import datetime

from django.http import HttpResponse

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ExportMixin:
    """
    Mixin genérico: añade exportación PDF y Excel a cualquier ListView.

    En la vista que lo incluya define:
        export_filename = 'nombre_sin_extension'
        export_fields   = [
            ('atributo',        'Encabezado'),   # campo simple
            ('relacion.campo',  'Encabezado'),   # dotted path
            (callable_fn,       'Encabezado'),   # función que recibe el objeto
        ]

    Para columnas dinámicas, sobreescribir get_dynamic_export_fields() en la vista.
    """

    export_fields: list = []
    export_filename: str = 'listado'

    # ── resolución de valor de celda ──────────────────────────────────────

    def _resolve_field(self, obj, attr_or_fn):
        if callable(attr_or_fn):
            return str(attr_or_fn(obj))
        val = obj
        for part in attr_or_fn.split('.'):
            val = getattr(val, part, '')
            if callable(val):
                val = val()
        return '' if val is None else str(val)

    def get_dynamic_export_fields(self):
        """
        Devuelve la lista de campos activos para exportar.
        Sobreescribir en subclases para columnas dinámicas.
        """
        return getattr(self, 'export_fields', [])

    def get_export_headers(self):
        return [label for _, label in self.get_dynamic_export_fields()]

    def get_export_row(self, obj):
        return [self._resolve_field(obj, attr) for attr, _ in self.get_dynamic_export_fields()]

    # ── PDF ───────────────────────────────────────────────────────────────

    def export_pdf(self, queryset):
        buffer = io.BytesIO()
        fields = self.get_dynamic_export_fields()
        num_cols = len(fields)

        # Orientación y tamaño de fuente inteligente según número de columnas
        if num_cols <= 3:
            page_size = A4
            font_hdr, font_data, pad = 11, 10, 6
        elif num_cols <= 5:
            page_size = landscape(A4)
            font_hdr, font_data, pad = 10, 9, 6
        elif num_cols <= 8:
            page_size = landscape(A4)
            font_hdr, font_data, pad = 9, 8, 5
        else:
            page_size = landscape(A4)
            font_hdr, font_data, pad = 8, 7, 4

        doc = SimpleDocTemplate(
            buffer, pagesize=page_size,
            leftMargin=1.5 * cm, rightMargin=1.5 * cm,
            topMargin=2 * cm, bottomMargin=1.5 * cm,
        )

        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(
            f'<b>Listado: {self.export_filename.capitalize()}</b>',
            styles['Title'],
        ))
        elements.append(Paragraph(
            f'Generado el {datetime.now().strftime("%d/%m/%Y a las %H:%M")}',
            styles['Normal'],
        ))
        elements.append(Spacer(1, 0.5 * cm))

        if not fields:
            elements.append(Paragraph('No hay columnas seleccionadas para exportar.', styles['Normal']))
            doc.build(elements)
            buffer.seek(0)
            fname = f'{self.export_filename}_{datetime.now().strftime("%Y%m%d")}.pdf'
            resp = HttpResponse(buffer.read(), content_type='application/pdf')
            resp['Content-Disposition'] = f'attachment; filename="{fname}"'
            return resp

        headers = [label for _, label in fields]
        rows = [[Paragraph(f'<b>{h}</b>', styles['Normal']) for h in headers]]
        for obj in queryset:
            rows.append([
                Paragraph(str(self._resolve_field(obj, attr)), styles['Normal'])
                for attr, _ in fields
            ])

        usable_width = page_size[0] - 3 * cm
        col_w = [usable_width / num_cols] * num_cols

        table = Table(rows, colWidths=col_w, repeatRows=1)
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND',      (0, 0), (-1, 0),  colors.HexColor('#1a56db')),
            ('TEXTCOLOR',       (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',        (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',        (0, 0), (-1, 0),  font_hdr),
            ('ALIGN',           (0, 0), (-1, 0),  'CENTER'),
            ('VALIGN',          (0, 0), (-1, 0),  'MIDDLE'),
            # Filas de datos
            ('FONTNAME',        (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE',        (0, 1), (-1, -1), font_data),
            ('ALIGN',           (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN',          (0, 1), (-1, -1), 'MIDDLE'),
            # Filas alternas
            ('ROWBACKGROUNDS',  (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            # Bordes y padding
            ('GRID',            (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
            ('TOPPADDING',      (0, 0), (-1, -1), pad),
            ('BOTTOMPADDING',   (0, 0), (-1, -1), pad),
            ('LEFTPADDING',     (0, 0), (-1, -1), pad),
            ('RIGHTPADDING',    (0, 0), (-1, -1), pad),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        fname = f'{self.export_filename}_{datetime.now().strftime("%Y%m%d")}.pdf'
        resp = HttpResponse(buffer.read(), content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        return resp

    # ── Excel ─────────────────────────────────────────────────────────────

    def export_excel(self, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.export_filename.capitalize()[:31]

        fields = self.get_dynamic_export_fields()
        headers = [label for _, label in fields]

        if not headers:
            ws.cell(row=1, column=1, value='No hay columnas seleccionadas.')
        else:
            hdr_fill  = PatternFill(fill_type='solid', fgColor='1a56db')
            hdr_font  = Font(bold=True, color='FFFFFF', size=10)
            hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin      = Side(style='thin', color='DEE2E6')
            border    = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col, text in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=text)
                cell.fill      = hdr_fill
                cell.font      = hdr_font
                cell.alignment = hdr_align
                cell.border    = border

            ws.row_dimensions[1].height = 20

            alt_fill   = PatternFill(fill_type='solid', fgColor='F8F9FA')
            data_align = Alignment(vertical='center', wrap_text=True)

            for row_idx, obj in enumerate(queryset, start=2):
                fill = alt_fill if row_idx % 2 == 0 else None
                for col_idx, (attr, _) in enumerate(fields, start=1):
                    value = self._resolve_field(obj, attr)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = data_align
                    cell.border    = border
                    if fill:
                        cell.fill = fill

            for col in ws.columns:
                max_len = max(
                    (len(str(c.value)) for c in col if c.value),
                    default=10,
                )
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

        ws.freeze_panes = 'A2'

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        fname = f'{self.export_filename}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        resp = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        return resp

    # ── interceptor de GET ────────────────────────────────────────────────

    def get(self, request, *args, **kwargs):
        fmt = request.GET.get('export')
        if fmt == 'pdf':
            return self.export_pdf(self.get_queryset())
        if fmt == 'excel':
            return self.export_excel(self.get_queryset())
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        params = self.request.GET.copy()
        params.pop('page', None)
        params.pop('export', None)
        ctx['query_params'] = params.urlencode()
        return ctx




